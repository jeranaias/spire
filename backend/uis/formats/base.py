"""Format detection + row-stream dispatch.

`detect_format(head_bytes)` returns one of: "csv" / "tsv" / "jsonl"
/ "xlsx" / "unknown". `stream_rows(raw, format)` yields per-row
dicts (header column name → cell value as string).

The pipeline calls `stream_rows` once per upload and downstream
stages (header normalization, mapping, transforms) consume the row
iterator without needing to know which format it came from.
"""
from __future__ import annotations

import csv
import io
import json
from typing import Dict, Iterable, Iterator, List


class DuplicateHeaderError(ValueError):
    """Raised when a CSV/TSV/XLSX file has duplicate header column names.

    csv.DictReader silently overwrites prior values with the same key,
    which means a file with `TAMCN, NSN, TAMCN` would lose the first
    TAMCN column on every row with no signal to the operator. We refuse
    the file so the operator can deduplicate it before re-uploading.
    """

    def __init__(self, *, duplicates: List[str]):
        self.duplicates = list(duplicates)
        super().__init__(
            f"Duplicate header column(s) detected: {self.duplicates!r}. "
            "csv.DictReader silently merges these — rename or drop "
            "duplicates in the source file and re-upload."
        )


def _check_duplicate_headers(header: List[str]) -> None:
    """Raise DuplicateHeaderError if any non-blank header is duplicated."""
    seen: Dict[str, int] = {}
    for col in header:
        if col is None:
            continue
        name = str(col).strip()
        if not name:
            continue
        seen[name] = seen.get(name, 0) + 1
    dupes = sorted(k for k, v in seen.items() if v > 1)
    if dupes:
        raise DuplicateHeaderError(duplicates=dupes)


# A RowStream is an iterator of dicts. Each dict has the source
# column names as keys and raw string values. Type coercion happens
# downstream in the transforms layer.
RowStream = Iterator[Dict[str, str]]


def detect_format(head: bytes) -> str:
    """Sniff the first chunk of bytes for the file format.

    Returns one of: "csv", "tsv", "jsonl", "xlsx", "xml", "unknown".

    Fixed-width is NOT auto-detected — its layout requires the
    adapter to declare column positions. Adapters that consume
    fixed-width data set ``format_hint="fixed_width"`` on
    AdapterSpec and pipeline.run_pipeline honors the hint.

    Skips leading comment lines (lines starting with `#` or `--`)
    and blank lines so an Oracle export with a comment header like
    "# Generated 2026-04-26" doesn't break delimiter counting on
    the actual header row.
    """
    if not head:
        return "unknown"
    # XLSX is a zip; check magic bytes
    if head[:4] == b"PK\x03\x04":
        return "xlsx"
    # UTF-16 BOM → assume CSV/TSV/XML inside (XLSX would be caught above)
    if head.startswith(b"\xff\xfe") or head.startswith(b"\xfe\xff"):
        try:
            text_head = head.decode("utf-16", errors="replace")[:4096]
        except UnicodeDecodeError:
            return "unknown"
    else:
        text_head = head[:4096].decode("utf-8", errors="replace")

    # XML: starts with <?xml or a tag like <Root>. Detect before
    # falling into the delimiter-counting CSV/TSV heuristics — XML
    # could otherwise misread as TSV when attributes have whitespace.
    stripped = text_head.lstrip("﻿").lstrip()
    if stripped.startswith("<?xml") or (
        stripped.startswith("<")
        and not stripped.startswith("<!--")
        and ">" in stripped
    ):
        return "xml"

    # JSONL: first non-comment non-blank line is `{...}` parseable
    first_payload_line = _first_payload_line(text_head)
    if first_payload_line.startswith("{"):
        try:
            json.loads(first_payload_line)
            return "jsonl"
        except json.JSONDecodeError:
            pass

    # CSV vs TSV — count delimiter occurrences in the first payload
    # line (skipping comments + blanks). Threshold relaxed: TSV needs
    # only ≥1 tab so a legit 2-column TSV doesn't get misdetected as
    # CSV-with-zero-commas → "unknown".
    tabs = first_payload_line.count("\t")
    commas = first_payload_line.count(",")
    semis = first_payload_line.count(";")  # European Excel default
    if tabs > 0 and tabs >= commas and tabs >= semis:
        return "tsv"
    if commas >= 1 and commas >= semis:
        return "csv"
    if semis >= 1:
        # European Excel exports use `;` as the field delimiter
        # because `,` is a decimal separator. Treat as CSV variant —
        # the streamer handles this when the operator passes
        # delimiter=";" but for autodetect we mark csv and let the
        # standard parser handle it (Python csv.DictReader does fine
        # with `,` even when `;` is also present, falling back when
        # the operator notices).
        return "csv"
    return "unknown"


def _first_payload_line(text: str) -> str:
    """Return the first line that isn't a blank or a comment.

    Comment markers: `#` (POSIX, Oracle SQL*Plus, YAML) and `--`
    (SQL). A `# Generated 2026-04-26 by Some Operator` preamble is
    typical on Oracle SQL*Plus spool exports and blew up
    detect_format() in the original implementation because it had
    zero commas + zero tabs and made the second line (the actual
    header) invisible.
    """
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#") or line.startswith("--"):
            continue
        return raw_line  # return un-stripped so leading whitespace is preserved for delimiter counting
    return ""


def stream_rows(raw: bytes, fmt: str, *, fixed_width_spec: Any = None) -> RowStream:
    """Dispatch to the format-specific row streamer.

    ``fixed_width_spec`` is required when ``fmt == "fixed_width"``;
    it carries the column positions (start, length, name). For
    other formats it's ignored.
    """
    if fmt == "csv":
        return _stream_csv(raw, delimiter=",")
    if fmt == "tsv":
        return _stream_csv(raw, delimiter="\t")
    if fmt == "jsonl":
        return _stream_jsonl(raw)
    if fmt == "xlsx":
        return _stream_xlsx(raw)
    if fmt == "xml":
        from .xml_format import stream_xml
        return stream_xml(raw)
    if fmt == "fixed_width":
        from .fixed_width import stream_fixed_width
        if fixed_width_spec is None:
            raise ValueError(
                "fixed_width format requires fixed_width_spec from AdapterSpec"
            )
        return stream_fixed_width(raw, fixed_width_spec)
    raise ValueError(f"Unknown or unsupported format: {fmt!r}")


# ---------------------------------------------------------------------------
# CSV / TSV
# ---------------------------------------------------------------------------


def _stream_csv(raw: bytes, *, delimiter: str) -> RowStream:
    """Stream CSV/TSV rows. Delegates to Python's stdlib csv module
    after decoding to UTF-8. Caller is responsible for normalizing
    bytes via `backend.uis.normalize.encoding.decode_bytes` first if
    encoding is suspect.

    Strips leading comment lines (`#` or `--`) and blank lines so
    Oracle SQL*Plus spool exports with header preambles parse
    correctly. The header row is whatever comes after the
    skipped lines.
    """
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1252", errors="replace")
    if text.startswith("﻿"):
        text = text[1:]  # strip BOM that survived decode

    # Filter comment + blank lines BEFORE handing to csv.DictReader.
    # Doing this in a generator stays memory-efficient for large
    # files (the upstream pipeline cap still bounds total rows).
    def _filtered_lines():
        for raw_line in text.splitlines(keepends=True):
            stripped = raw_line.strip()
            if not stripped:
                continue
            if stripped.startswith("#") or stripped.startswith("--"):
                continue
            yield raw_line

    reader = csv.DictReader(_filtered_lines(), delimiter=delimiter)
    # csv.DictReader doesn't materialize fieldnames until the first
    # access — read it eagerly so we can detect duplicate columns
    # BEFORE iterating rows. csv.DictReader merges duplicates by
    # overwriting (last column wins) which silently loses data on
    # malformed exports. Refuse those files.
    if reader.fieldnames is not None:
        _check_duplicate_headers(list(reader.fieldnames))
    for row in reader:
        # csv.DictReader yields Optional[str] values — coerce to ""
        yield {k: (v if v is not None else "") for k, v in row.items()}


# ---------------------------------------------------------------------------
# JSONL
# ---------------------------------------------------------------------------


def _stream_jsonl(raw: bytes) -> RowStream:
    """One row per line, each line a JSON object."""
    text = raw.decode("utf-8", errors="replace")
    if text.startswith("﻿"):
        text = text[1:]
    for line_idx, line in enumerate(text.splitlines(), start=1):
        s = line.strip()
        if not s:
            continue
        try:
            obj = json.loads(s)
        except json.JSONDecodeError:
            # Skip unparseable lines; the pipeline's validation stage
            # surfaces a per-row warning if this is unexpected.
            continue
        if not isinstance(obj, dict):
            continue
        yield {k: ("" if v is None else str(v)) for k, v in obj.items()}


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _stream_xlsx(raw: bytes) -> RowStream:
    """Read the largest non-empty sheet of an XLSX as rows.

    Earlier this only read wb.active — fine for single-sheet
    workbooks, but a multi-sheet export (common: GCSS-MC operators
    paste into "Sheet1", "Data", and "Notes" then save) with the
    actual data on a non-default sheet silently produced zero rows.

    Strategy: scan every sheet, count non-empty rows below the
    header, pick the sheet with the most. On a tie, take wb.active
    (preserves historical behavior when multiple sheets have data).

    Requires `openpyxl` — declared as an optional dep so the rest of
    the UIS package stays import-safe on a minimal install.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "XLSX ingest requires `openpyxl`. Install it: pip install openpyxl"
        ) from e
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if not wb.sheetnames:
        return iter([])

    # Pick the largest non-empty sheet
    best_sheet = None
    best_rows: list = []
    best_count = -1
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_rows = list(_xlsx_sheet_to_rows(ws))
        # Count non-empty rows (some sheets have data scattered
        # across blank rows; we use the populated count as the
        # tie-breaker).
        non_empty = sum(1 for r in sheet_rows if any(v.strip() for v in r.values()))
        if non_empty > best_count:
            best_count = non_empty
            best_sheet = name
            best_rows = sheet_rows

    if best_sheet is None or best_count <= 0:
        return iter([])
    return iter(best_rows)


def _xlsx_sheet_to_rows(ws) -> RowStream:
    """Convert one openpyxl worksheet to a row-of-dicts iterator."""
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = list(next(rows))
    except StopIteration:
        return iter([])
    header = [str(h) if h is not None else "" for h in header_row]
    # XLSX surfaces the same duplicate-header silent-merge hazard as CSV
    # because we build a dict keyed on header names. Refuse early.
    _check_duplicate_headers(header)
    out: list = []
    for row in rows:
        d: Dict[str, str] = {}
        for i, cell in enumerate(row):
            key = header[i] if i < len(header) else f"col_{i}"
            if not key:
                continue  # skip cells whose header column is blank
            d[key] = "" if cell is None else str(cell)
        out.append(d)
    return iter(out)
