"""
XLSX export layer -- formats every output to look like it was exported from
GCSS-MC / DRRS-MC: navy header bar, alternating row shading, frozen header,
auto-filter, condition-based cell coloring.

All six output files land in dataset/data/out/:
  gcss_mc_mpr_export.xlsx         (SENTRY / PULSE input)
  daily_readiness_snapshots.xlsx  (PULSE input)
  installation_incident_log.xlsx  (BASTION input)
  fleet_registry.xlsx             (ground truth)
  parts_catalog.xlsx              (ground truth)
  personnel_roster.xlsx           (ground truth -- never leaves local drive)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent / "data" / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Styling tokens
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1B365D")        # navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F6FA")       # light blue
BORDER = Border(
    left=Side(style="thin", color="CDD3DE"),
    right=Side(style="thin", color="CDD3DE"),
    top=Side(style="thin", color="CDD3DE"),
    bottom=Side(style="thin", color="CDD3DE"),
)

READINESS_FILLS = {
    "MC":    PatternFill("solid", fgColor="C8E6C9"),  # green
    "PMC":   PatternFill("solid", fgColor="FFF9C4"),  # yellow
    "NMCM":  PatternFill("solid", fgColor="FFCCBC"),  # orange
    "NMCS":  PatternFill("solid", fgColor="FFCDD2"),  # red
}

CLASSIFICATION_FILLS = {
    "UNCLASSIFIED": PatternFill("solid", fgColor="C8E6C9"),
    "CUI":          PatternFill("solid", fgColor="FFF9C4"),
    "CONFIDENTIAL": PatternFill("solid", fgColor="FFCCBC"),
    "SECRET":       PatternFill("solid", fgColor="FFCDD2"),
}

SEVERITY_FILLS = {
    "LOW":      PatternFill("solid", fgColor="C8E6C9"),
    "MODERATE": PatternFill("solid", fgColor="FFF9C4"),
    "HIGH":     PatternFill("solid", fgColor="FFCCBC"),
    "CRITICAL": PatternFill("solid", fgColor="FFCDD2"),
}

CONDITION_FILLS = {
    "Deadlined": PatternFill("solid", fgColor="FFCDD2"),
    "Degraded":  PatternFill("solid", fgColor="FFF9C4"),
    "Minor":     PatternFill("solid", fgColor="C8E6C9"),
}

BOOL_TRUE_FILL = PatternFill("solid", fgColor="FFCDD2")


# ---------------------------------------------------------------------------
# Sheet-writing helpers
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, sheet_name: str, columns: list, rows: Iterable[list], *, column_color_rules: dict | None = None) -> None:
    """Append a new sheet with the standard SPIRE formatting.

    column_color_rules: { column_index_1based: dict-of-value-to-fill } -- each
    cell in that column gets the matching fill, if present.
    """
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
    ws.append(columns)

    # Header styling
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Data rows
    row_idx = 1
    for row in rows:
        row_idx += 1
        ws.append(row)
        alt = (row_idx % 2 == 0)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            if alt:
                cell.fill = ALT_ROW_FILL
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Condition-based coloring overrides stripe
            if column_color_rules and col_idx in column_color_rules:
                rule = column_color_rules[col_idx]
                match = rule.get(cell.value)
                if match is not None:
                    cell.fill = match

    # Column widths: fit content (clamped 10-60)
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = len(str(columns[col_idx - 1]))
        for cell in col[:1000]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = "A2"
    # Auto-filter on header row
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{row_idx}"


def _drop_default_sheet(wb: Workbook) -> None:
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def export_mpr(srs, path: Path = OUT_DIR / "gcss_mc_mpr_export.xlsx") -> Path:
    wb = Workbook()
    _drop_default_sheet(wb)

    columns = [
        "SR Number", "Open Date", "Close Date", "Unit UIC", "Unit", "Asset ID",
        "Equipment Type", "TAMCN", "NSN", "Serial Number",
        "Job Status", "Condition", "Priority",
        "Defect Code Primary", "Defect Code Secondary", "Fault Component", "Fault ID",
        "Maintenance Level", "TM Reference",
        "Source Classification", "Detected Classification", "Sensitive Flags",
        "Mechanic EDIPI", "Mechanic Name",
        "Labor Hours Est", "Labor Hours Actual",
        "Parts Cost Est", "Parts Cost Actual",
        "Evacuated To", "Data Quality Flag",
        "Is PMCS", "Remark",
    ]

    rows = []
    for sr in srs:
        rows.append([
            sr.sr_number, sr.open_date.isoformat(),
            sr.close_date.isoformat() if sr.close_date else "",
            sr.unit_uic, sr.unit_name, sr.asset_id,
            sr.equipment_type, sr.tamcn, sr.nsn, sr.serial_number,
            sr.job_status, sr.condition, sr.priority,
            sr.defect_code_primary, sr.defect_code_secondary, sr.fault_component, sr.fault_id,
            sr.maintenance_level, sr.tm_reference,
            sr.source_classification, sr.detected_classification, ", ".join(sr.sensitive_flags),
            sr.mechanic_edipi, sr.mechanic_name,
            sr.labor_hours_est, sr.labor_hours_actual,
            sr.parts_cost_est, sr.parts_cost_actual,
            sr.evacuated_to, sr.data_quality_flag,
            "YES" if sr.is_pmcs else "NO", sr.remark_text,
        ])

    _write_sheet(wb, "MPR Export", columns, rows, column_color_rules={
        12: CONDITION_FILLS,
        20: CLASSIFICATION_FILLS,
        21: CLASSIFICATION_FILLS,
    })
    wb.save(path)
    return path


def export_daily_readiness(snapshots, path: Path = OUT_DIR / "daily_readiness_snapshots.xlsx") -> Path:
    """Use openpyxl write-only mode -- 127K rows with full formatting is ~60x
    faster this way (avoids the regular-workbook cell-object overhead)."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Daily Readiness")

    columns = [
        "Snapshot Date", "Asset ID", "Unit UIC", "Unit", "Equipment Type", "TAMCN", "Serial",
        "Readiness", "Condition", "Open SR Count", "Days Deadlined", "Days Since Maint",
        "Current Hours", "Current Miles", "Parts On Order", "Location", "Deployment Status",
    ]
    from openpyxl.cell import WriteOnlyCell
    header_cells = []
    for col in columns:
        c = WriteOnlyCell(ws, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        header_cells.append(c)
    ws.append(header_cells)

    for s in snapshots:
        readiness_cell = WriteOnlyCell(ws, value=s.readiness_code)
        readiness_cell.fill = READINESS_FILLS.get(s.readiness_code, None) or PatternFill()
        cond_cell = WriteOnlyCell(ws, value=s.condition)
        cond_cell.fill = CONDITION_FILLS.get(s.condition, None) or PatternFill()
        ws.append([
            s.snapshot_date.isoformat(), s.asset_id, s.unit_uic, s.unit_name,
            s.equipment_type, s.tamcn, s.serial_number,
            readiness_cell, cond_cell,
            s.open_sr_count, s.days_deadlined, s.days_since_maintenance,
            s.current_hours, s.current_miles, s.parts_on_order, s.location, s.deployment_status,
        ])

    # Reasonable column widths (write-only doesn't auto-fit, so preset).
    widths = [12, 24, 10, 14, 16, 8, 18, 10, 10, 10, 10, 10, 14, 14, 12, 22, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    return path


def export_incidents(incidents, path: Path = OUT_DIR / "installation_incident_log.xlsx") -> Path:
    wb = Workbook()
    _drop_default_sheet(wb)
    columns = [
        "Incident #", "Date/Time", "Type", "Severity",
        "Location Building", "Location Grid", "Location Description",
        "Initial Report", "FPCON at Time", "FPCON Change",
        "Response Force", "Response Time (min)",
        "Actions Taken", "Casualties", "Property Damage USD",
        "Resolution", "Lessons Learned", "Reported By", "Watch Officer",
    ]
    rows = []
    for i in incidents:
        rows.append([
            i.incident_number, i.date_time.isoformat(), i.type, i.severity,
            i.location_building, i.location_grid, i.location_description,
            i.initial_report, i.fpcon_at_time, i.fpcon_change or "",
            i.response_force, i.response_time_minutes,
            i.actions_taken, i.casualties, i.property_damage_usd,
            i.resolution, i.lessons_learned, i.reported_by, i.watch_officer,
        ])
    _write_sheet(wb, "Incident Log", columns, rows, column_color_rules={
        4: SEVERITY_FILLS,
    })
    wb.save(path)
    return path


def export_fleet_registry(assets, path: Path = OUT_DIR / "fleet_registry.xlsx") -> Path:
    wb = Workbook()
    _drop_default_sheet(wb)
    columns = [
        "Asset ID", "Unit UIC", "Unit", "Parent Command", "Location",
        "Equipment Type", "Nomenclature", "Model", "TAMCN", "NSN", "Serial Number", "FSC",
        "Fielding Date", "Initial Hours", "Initial Miles",
        "Optempo", "Deployment Status", "Classification Risk",
    ]
    rows = []
    for a in assets:
        rows.append([
            a.asset_id, a.unit_uic, a.unit_name, a.unit_parent, a.location,
            a.equipment_type, a.nomenclature, a.model, a.tamcn, a.nsn, a.serial_number, a.fsc,
            a.fielding_date.isoformat(), a.initial_hours, a.initial_miles,
            a.optempo, a.deployment_status, a.classification_risk,
        ])
    _write_sheet(wb, "Fleet Registry", columns, rows)
    wb.save(path)
    return path


def export_personnel_roster(roster, path: Path = OUT_DIR / "personnel_roster.xlsx") -> Path:
    wb = Workbook()
    _drop_default_sheet(wb)
    columns = [
        "EDIPI", "Rank", "Last Name", "First Initial", "MOS", "Unit", "Phone Ext", "SSN Last 4", "Email",
    ]
    rows = []
    for m in roster:
        rows.append([
            m.edipi, m.rank, m.last_name, m.first_initial, m.mos, m.unit_name, m.phone_ext, m.ssn_last4, m.email,
        ])
    _write_sheet(wb, "Personnel Roster", columns, rows)
    wb.save(path)
    return path


def export_parts_catalog(requisitions, path: Path = OUT_DIR / "parts_catalog.xlsx") -> Path:
    """Build a de-duplicated parts catalog from every requisition's NSN / nomenclature."""
    wb = Workbook()
    _drop_default_sheet(wb)
    seen: dict = {}
    for req in requisitions:
        if req.nsn in seen:
            seen[req.nsn]["orders"] += 1
            continue
        seen[req.nsn] = {
            "nsn": req.nsn,
            "nomenclature": req.nomenclature,
            "unit_cost": req.unit_cost,
            "uoi": req.uoi,
            "orders": 1,
            "supply_path": req.supply_path,
        }
    columns = ["NSN", "Nomenclature", "Unit Cost USD", "UOI", "Typical Supply Path", "Times Ordered"]
    rows = []
    for entry in seen.values():
        rows.append([
            entry["nsn"], entry["nomenclature"], entry["unit_cost"], entry["uoi"],
            entry["supply_path"], entry["orders"],
        ])
    _write_sheet(wb, "Parts Catalog", columns, rows)
    wb.save(path)
    return path


def export_metadata(run_stats: dict, path: Path = OUT_DIR / "dataset_metadata.xlsx") -> Path:
    wb = Workbook()
    _drop_default_sheet(wb)
    columns = ["Key", "Value"]
    rows = [[k, str(v)] for k, v in run_stats.items()]
    _write_sheet(wb, "Dataset Metadata", columns, rows)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# One-call export
# ---------------------------------------------------------------------------

def export_all(*, units, assets, roster, srs, snapshots, reqs, incidents, cannib_events, violations, run_stats) -> dict:
    paths = {
        "mpr":         export_mpr(srs),
        "readiness":   export_daily_readiness(snapshots),
        "incidents":   export_incidents(incidents),
        "fleet":       export_fleet_registry(assets),
        "personnel":   export_personnel_roster(roster),
        "parts":       export_parts_catalog(reqs),
        "metadata":    export_metadata(run_stats),
    }
    return paths


if __name__ == "__main__":
    # Minimal smoke test -- generate a tiny dataset and export.
    from config import RANDOM_SEED, OUTPUT_TARGETS
    from fleet import generate_fleet
    from personnel import generate_personnel
    from lifecycle import run_simulation
    from consistency import inject_cannibalizations, inject_data_quality_defects, run_all_checks
    from incidents import generate_incidents

    units, assets = generate_fleet(RANDOM_SEED)
    roster = generate_personnel(units, OUTPUT_TARGETS["personnel_count"], RANDOM_SEED)
    srs, snaps, reqs = run_simulation(units, assets, roster, RANDOM_SEED)
    inject_data_quality_defects(srs, snaps, RANDOM_SEED)
    cans = inject_cannibalizations(srs, assets, RANDOM_SEED)
    violations = run_all_checks(srs, assets, snaps, cans)
    incidents = generate_incidents(RANDOM_SEED)

    run_stats = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "random_seed": RANDOM_SEED,
        "units": len(units),
        "assets": len(assets),
        "personnel": len(roster),
        "service_requests": len(srs),
        "daily_snapshots": len(snaps),
        "part_requisitions": len(reqs),
        "cannibalization_events": len(cans),
        "incidents": len(incidents),
        "consistency_errors": sum(1 for v in violations if v.severity == "error"),
        "consistency_warnings": sum(1 for v in violations if v.severity == "warning"),
    }

    paths = export_all(
        units=units, assets=assets, roster=roster,
        srs=srs, snapshots=snaps, reqs=reqs,
        incidents=incidents, cannib_events=cans,
        violations=violations, run_stats=run_stats,
    )
    for name, p in paths.items():
        size = p.stat().st_size / 1024
        print(f"  {name:<12} -> {p.name}  ({size:,.1f} KB)")
